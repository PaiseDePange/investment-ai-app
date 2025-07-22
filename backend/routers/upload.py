from fastapi import APIRouter, UploadFile, File, HTTPException
from fastapi.responses import JSONResponse
import pandas as pd
import io
from openpyxl import load_workbook
from collections import Counter

router = APIRouter()

def format_column_headers(headers):
    formatted = []
    blank_counter = 1
    for h in headers:
        try:
            h_parsed = pd.to_datetime(h)
            formatted.append(h_parsed.strftime("%b-%Y"))
        except:
            if pd.notnull(h) and str(h).strip():
                formatted.append(str(h))
            else:
                formatted.append(f"Unnamed_{blank_counter}")
                blank_counter += 1
    counts = Counter()
    unique = []
    for h in formatted:
        counts[h] += 1
        unique.append(f"{h}_{counts[h]}" if counts[h] > 1 else h)
    return unique

def extract_table(df, start_label, start_row_offset, col_count=11):
    start_row = df[df.iloc[:, 0] == start_label].index[0]
    header_row = start_row + start_row_offset
    headers_raw = df.iloc[header_row, 0:col_count].tolist()
    headers = format_column_headers(headers_raw)
    data_rows = []
    for i in range(start_row + start_row_offset + 1, df.shape[0]):
        row = df.iloc[i, 0:col_count]
        if row.isnull().all():
            break
        data_rows.append(row.tolist())
    df_temp = pd.DataFrame(data_rows, columns=headers)
    df_temp = df_temp.loc[:, df_temp.iloc[0].notna()]
    df_temp.fillna(0, inplace=True)
    return df_temp

@router.post("/upload-excel")
async def upload_excel(file: UploadFile = File(...)):
    contents = await file.read()

    wb = load_workbook(io.BytesIO(contents), data_only=True)
    ws = wb["Data Sheet"]
    df_all = pd.DataFrame(ws.values)

    try:
        df_meta = extract_table(df_all, "META", 1, 2)
        df_meta.columns = ["Label", "Value"]
        df_meta = df_meta.set_index("Label")
        
        company_name = ws["B1"].value or "Could not find Company Name"

        def safe_get(label):
            try:
                return float(df_meta.loc[label, "Value"])
            except:
                return 0.0

        current_price = safe_get("Current Price")
        market_cap = safe_get("Market Capitalization")

        df_bs = extract_table(df_all, "BALANCE SHEET", 1).set_index("Report Date")
        df_pl = extract_table(df_all, "PROFIT & LOSS", 1).set_index("Report Date")

        revenue_row = df_pl.loc["Sales"].dropna()
        tax_row = df_pl.loc["Tax"].dropna()
        depreciation_row = df_pl.loc["Depreciation"].dropna()
        
        #share_outstanding_row = df_bs.loc["No. of Equity Shares"].dropna() if "No. of Equity Shares" in df_bs.index else pd.Series([0])
        last = share_outstanding_row[-1]
        second_last = share_outstanding_row.iloc[-2] if len(share_outstanding_row) > 1 else 0
        shares_outstanding = round((second_last if last == 0 else last) / 1e7, 2)
        
        calculated_ebit = revenue_row[-1] - sum(df_pl.loc[row].dropna()[-1] for row in [
            "Raw Material Cost", "Change in Inventory", "Power and Fuel",
            "Other Mfr. Exp", "Employee Cost", "Selling and admin", "Other Expenses"
        ] if row in df_pl.index)
        latest_revenue = revenue_row[-1]

        ebit_margin = round((calculated_ebit / latest_revenue) * 100, 1) if latest_revenue else 0
        tax_rate = round((tax_row[-1]/calculated_ebit)*100, 1) if calculated_ebit else 0
        depreciation_pct = round((depreciation_row[-1]/latest_revenue)*100, 1) if latest_revenue else 0

        shares_outstanding = round(share_outstanding_row[-1]/1e7, 2) if not share_outstanding_row.empty else 0

        debt = float(df_bs.loc["Borrowings"].dropna()[-1]) if "Borrowings" in df_bs.index else 0
        investments = float(df_bs.loc["Investments"].dropna()[-1]) if "Investments" in df_bs.index else 0
        cash = float(df_bs.loc["Cash & Bank"].dropna()[-1]) if "Cash & Bank" in df_bs.index else 0
        net_debt = debt - (investments + cash)

        missing_keys = []
        if current_price == 0: missing_keys.append("Current Price")
        if shares_outstanding == 0: missing_keys.append("Shares Outstanding")
        if latest_revenue == 0: missing_keys.append("Sales")

        assumptions = {
            "ebit_margin": ebit_margin,
            "depreciation_pct": depreciation_pct,
            "tax_rate": tax_rate,
            "capex_pct": 2.0,
            "wc_change_pct": 2.0,
            "interest_pct": 11.0,
            "growth_x": 20.0,
            "growth_y": 12.0,
            "growth_terminal": 4.0,
            "period_x": 5,
            "period_y": 15,
            "shares_outstanding": shares_outstanding,
            "net_debt": net_debt,
            "latest_revenue": latest_revenue,
            "current_price": current_price,
            "missing_fields": missing_keys,
            "company_name": company_name
        }
        for key, val in assumptions.items():
            if isinstance(val, float):
                assumptions[key] = round(val, 2)

        return JSONResponse(content=assumptions)

    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=400)
